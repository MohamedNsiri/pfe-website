import xml.etree.ElementTree as ET
import csv
from io import StringIO
import openpyxl
from openpyxl.utils import get_column_letter
import re
from fpdf import FPDF
import os
from datetime import datetime
import spacy

class SBOMValidator:
    def __init__(self, xml_file_path=None, excel_file_path=None):

        self.xml_data = None
        self.excel_data = None
        self.nlp = spacy.load("en_core_web_sm")

        wcpr = ""
        wcpar = ""
        wcusfa = ""
        
        if xml_file_path:
            self._parse_xml(xml_file_path)
            
        if excel_file_path:
            self._parse_excel(excel_file_path)
            
    def _parse_xml(self, file_path):
        self.xml_data = {
            'sboms': [],
            'file_type': 'xml'
        }        
        tree = ET.parse(file_path)
        root = tree.getroot()        
        for sbom in root.findall('sbom'):
            sbom_data = {
                'attributes': sbom.attrib,
                'subassemblies': [],
                'cost_results': [],
                'bom_elements': []
                #'general_special_wire_pmd' : []
            }            
            for sub in sbom.findall('sbomsubassembly'):
                sbom_data['subassemblies'].append({
                    'attributes': sub.attrib,
                    'parent_id': sub.attrib.get('parentsubid', None)
                })            
            for cost in sbom.findall('.//costresult'):
                sbom_data['cost_results'].append(cost.attrib)            
            for bom in sbom.findall('.//bomelement'):
                sbom_data['bom_elements'].append(bom.attrib)
            #for pmd in sbom.findall('.//GeneralSpecialWirePMD'):
                #sbom_data['general_special_wire_pmd'].append(pmd.attrib)
            
            self.xml_data['sboms'].append(sbom_data)

    def _parse_excel(self, file_path):
        self.excel_data = {
            'sheets': [],
            'workbook': None,
            'file_type': 'excel'
        }        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.excel_data['workbook'] = workbook
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'index': workbook.index(sheet),
                'data': [],
                'headers': [],
                'empty': True
            }            
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                sheet_data['empty'] = False                
                if rows[0]:
                    sheet_data['headers'] = list(rows[0])                
                for row in rows[1:]:
                    if any(cell is not None for cell in row):
                        sheet_data['data'].append(row)            
            self.excel_data['sheets'].append(sheet_data)

    # XML-specific methods
    def get_sbom_attributes(self):
        if not self.xml_data:
            raise ValueError("No XML data loaded")
        return [sbom['attributes'] for sbom in self.xml_data['sboms']]

    def get_subassemblies(self, flatten_attributes=False):
        if not self.xml_data:
            raise ValueError("No XML data loaded")
        subassemblies = []
        for sbom in self.xml_data['sboms']:
            for sub in sbom['subassemblies']:
                if flatten_attributes:
                    sub_data = sub['attributes'].copy()
                    sub_data['parent_id'] = sub['parent_id']
                    subassemblies.append(sub_data)
                else:
                    subassemblies.append(sub)
        return subassemblies

    def get_cost_results(self, filter=None):
        if not self.xml_data:
            raise ValueError("No XML data loaded")

        cost_results = []
        for sbom in self.xml_data['sboms']:
            for result in sbom.get('cost_results', []):
                if filter:
                    cost_results.append(result.get(filter, ""))
                else:
                    cost_results.append(result)
        return cost_results

    def get_bom_elements(self):
        """Get all BOM elements."""
        if not self.xml_data:
            raise ValueError("No XML data loaded")
        bom_elements = []
        for sbom in self.xml_data['sboms']:
            bom_elements.extend(sbom['bom_elements'])
        return bom_elements

    # Excel-specific methods
    def get_sheet_names(self):
        """Get list of all sheet names in the Excel file."""
        if not self.excel_data:
            raise ValueError("No Excel data loaded")
        return [sheet['name'] for sheet in self.excel_data['sheets']]

    def get_sheet_by_name(self, sheet_name):
        """Get data from a specific sheet by name."""
        if not self.excel_data:
            raise ValueError("No Excel data loaded")
        
        for sheet in self.excel_data['sheets']:
            if sheet['name'] == sheet_name:
                return sheet
        return None

    def filter_sheets(self, index=None, name=None, filter_column=None, filter_value=None, return_column=None):
        if not self.excel_data:
            raise ValueError("No Excel data loaded")
        
        if index is None and name is None:
            raise ValueError("Either index or name must be provided")
        if index is not None and name is not None:
            raise ValueError("Only one of index or name should be provided")
        
        sheet = None
        if index is not None:
            if 0 <= index < len(self.excel_data['sheets']):
                sheet = self.excel_data['sheets'][index]
        else:
            for s in self.excel_data['sheets']:
                if s['name'] == name:
                    sheet = s
                    break
        
        if sheet is None:
            raise ValueError(f"Sheet not found (index={index}, name={name})")
        
        if filter_column is None and return_column is None:
            return sheet
            
        headers = sheet['headers']
        try:
            if filter_column:
                filter_col_idx = headers.index(filter_column)
            if return_column:
                return_col_idx = headers.index(return_column)
        except ValueError:
            raise ValueError(f"Column not found in sheet headers")
            
        results = []
        for row in sheet['data']:
            if filter_column is None or row[filter_col_idx] == filter_value:
                if return_column is None:
                    results.append(row)
                else:
                    results.append(row[return_col_idx])
                    
        return results if len(results) > 1 else results[0] if results else None

    def validate(self):
        if not self.xml_data:
            raise ValueError("No XML data loaded for validation")
        if not self.excel_data:
            raise ValueError("No Excel data loaded for validation")

        # Initialize validation results
        validation_results = {
            "status": "success",
            "message": "All values matched.",
            "mismatches": [],
            "workcenter_validation": [],
            "wire_length_validation": []
        }

        # ====================================================================
        # 1. Validate work center attributes from XML against input values
        # ====================================================================
        try:
            # Get the SBOM attributes from the XML data structure
            sbom_attributes = self.get_sbom_attributes()
            if not sbom_attributes:
                raise ValueError("No SBOM attributes found in XML data")
            
            # We'll use the first SBOM's attributes for validation
            attributes = sbom_attributes[0]
            
            # Get the work center attributes
            xml_wcpr = attributes.get('workcenterplantreference')
            xml_wcpar = attributes.get('workcenterproductionareareference')
            xml_wcusfa = attributes.get('workcenter_usesinglefinalassembly')        

            # Check if we have input values to compare against
            if hasattr(self, 'wcpr'):
                if xml_wcpr is None:
                    validation_results["workcenter_validation"].append(
                        "Workcenter Plant Reference missing in XML"
                    )
                elif str(self.wcpr) != str(xml_wcpr):
                    validation_results["workcenter_validation"].append(
                        f"Workcenter Plant Reference mismatch: Input={self.wcpr}, XML={xml_wcpr}"
                    )
            else:
                validation_results["workcenter_validation"].append(
                    "No Workcenter Plant Reference input provided for validation"
                )

            if hasattr(self, 'wcpar'):
                if xml_wcpar is None:
                    validation_results["workcenter_validation"].append(
                        "Workcenter Production Area Reference missing in XML"
                    )
                elif str(self.wcpar) != str(xml_wcpar):
                    validation_results["workcenter_validation"].append(
                        f"Workcenter Production Area Reference mismatch: Input={self.wcpar}, XML={xml_wcpar}"
                    )
            else:
                validation_results["workcenter_validation"].append(
                    "No Workcenter Production Area Reference input provided for validation"
                )

            if hasattr(self, 'wcusfa'):
                if xml_wcusfa is None:
                    validation_results["workcenter_validation"].append(
                        "Workcenter Use Single Final Assembly missing in XML"
                    )
                else:
                    # Handle case sensitivity and potential typo in attribute name
                    input_value = str(self.wcusfa).strip().lower()
                    xml_value = str(xml_wcusfa).strip().lower()
                    
                    if input_value != xml_value:
                        validation_results["workcenter_validation"].append(
                            f"Workcenter Use Single Final Assembly mismatch: Input={self.wcusfa}, XML={xml_wcusfa}"
                        )
            else:
                validation_results["workcenter_validation"].append(
                    "No Workcenter Use Single Final Assembly input provided for validation"
                )

        except Exception as e:
            validation_results["workcenter_validation"].append(
                f"Error validating work center attributes: {str(e)}"
            )

        # ====================================================================
        # 2. Validate twisted wires data against Excel
        # ====================================================================
        try:
            # Get all cost results with descriptions
            cost_results = self.get_cost_results()
            if not cost_results:
                validation_results["status"] = "error"
                validation_results["message"] = "No cost results found in SBOM XML."
                return validation_results

            # Now validate against Excel sheet "Twisted Wires"
            twisted_sheet = self.get_sheet_by_name("Twisted Wires")
            if not twisted_sheet:
                validation_results["status"] = "error"
                validation_results["message"] = "'Twisted Wires' sheet not found in Excel."
                return validation_results

            headers = twisted_sheet["headers"]
            data_rows = twisted_sheet["data"]

            # Get column indices for all relevant fields
            try:
                wire_nr_idx = headers.index("Wires Nr") if "Wires Nr" in headers else None
                pitch_idx = headers.index("Pitch") if "Pitch" in headers else None
                open_end1_idx = headers.index("Open end Length 1") if "Open end Length 1" in headers else None
                open_end2_idx = headers.index("Open end Length 2") if "Open end Length 2" in headers else None
                twist_len_idx = headers.index("Length of twist") if "Length of twist" in headers else None

                if None in [wire_nr_idx, pitch_idx, open_end1_idx, open_end2_idx, twist_len_idx]:
                    missing = [h for h, idx in zip(["Wires Nr", "Pitch", "Open end Length 1", 
                                                "Open end Length 2", "Length of twist"], 
                                                [wire_nr_idx, pitch_idx, open_end1_idx, 
                                                open_end2_idx, twist_len_idx]) if idx is None]
                    validation_results["status"] = "error"
                    validation_results["message"] = f"Required columns not found in Excel: {', '.join(missing)}"
                    return validation_results

                # Create mapping of wire numbers to all relevant values from Excel
                excel_wire_data = {}
                for row in data_rows:
                    wire_nr = row[wire_nr_idx]
                    if wire_nr:
                        excel_wire_data[wire_nr.strip()] = {
                            "pitch": row[pitch_idx],
                            "open_end1": row[open_end1_idx],
                            "open_end2": row[open_end2_idx],
                            "twist_len": row[twist_len_idx]
                        }

                for result in cost_results:
                    description = result.get("description", "")
                    if not description:
                        continue
                        
                    # Extract all relevant values from XML description
                    pitch_match = re.search(r"Pitch:\s*([\d.]+)", description)
                    untwist_a_match = re.search(r"Untwist A:\s*([\d.]+)", description)
                    untwist_b_match = re.search(r"Untwist B:\s*([\d.]+)", description)
                    twist_len_match = re.search(r"Twist length:\s*([\d.]+)", description)

                    if not all([pitch_match, untwist_a_match, untwist_b_match, twist_len_match]):
                        continue
                        
                    xml_values = {
                        "pitch": float(pitch_match.group(1)),
                        "open_end1": float(untwist_a_match.group(1)),
                        "open_end2": float(untwist_b_match.group(1)),
                        "twist_len": float(twist_len_match.group(1))
                    }
                    
                    # Extract wire pairs from the beginning of description
                    wire_matches = re.findall(r"Twist\s+([^,]+(?:,[^,]+)*)", description)
                    if not wire_matches:
                        continue
                        
                    # Get all wire numbers mentioned in this specific cost result
                    wires_in_result = []
                    for group in wire_matches:
                        wires_in_result.extend(re.findall(r"(\d+\(\d+\))", group))
                        
                    # Validate each wire in this cost result against Excel
                    for wire in wires_in_result:
                        wire_id = wire.strip()
                        if wire_id not in excel_wire_data:
                            validation_results["mismatches"].append(f"{wire_id} missing in Excel")
                            continue
                            
                        excel_data = excel_wire_data[wire_id]
                        
                        # Check each value
                        try:
                            # Check pitch
                            excel_pitch = float(excel_data["pitch"])
                            if excel_pitch != xml_values["pitch"]:
                                validation_results["mismatches"].append(
                                    f"Pitch mismatch for {wire_id}: "
                                    f"SBOM={xml_values['pitch']}, Excel={excel_pitch}"
                                )
                            
                            # Check open end length 1 (Untwist A)
                            excel_open_end1 = float(excel_data["open_end1"])
                            if excel_open_end1 != xml_values["open_end1"]:
                                validation_results["mismatches"].append(
                                    f"Open end Length 1 (Untwist A) mismatch for {wire_id}: "
                                    f"SBOM={xml_values['open_end1']}, Excel={excel_open_end1}"
                                )
                            
                            # Check open end length 2 (Untwist B)
                            excel_open_end2 = float(excel_data["open_end2"])
                            if excel_open_end2 != xml_values["open_end2"]:
                                validation_results["mismatches"].append(
                                    f"Open end Length 2 (Untwist B) mismatch for {wire_id}: "
                                    f"SBOM={xml_values['open_end2']}, Excel={excel_open_end2}"
                                )
                            
                            # Check length of twist
                            excel_twist_len = float(excel_data["twist_len"])
                            if excel_twist_len != xml_values["twist_len"]:
                                validation_results["mismatches"].append(
                                    f"Length of twist mismatch for {wire_id}: "
                                    f"SBOM={xml_values['twist_len']}, Excel={excel_twist_len}"
                                )
                                
                        except (ValueError, TypeError) as e:
                            validation_results["mismatches"].append(f"Invalid numeric format in Excel for {wire_id}: {str(e)}")

            except Exception as e:
                validation_results["status"] = "error"
                validation_results["message"] = f"Error during twisted wires validation: {str(e)}"
                return validation_results

        except Exception as e:
            validation_results["status"] = "error"
            validation_results["message"] = f"Error during validation: {str(e)}"
            return validation_results

        # ====================================================================
        # 3. Validate wire lengths from XML subassemblies against Excel
        # ====================================================================
        try:
            # Get the "Wires Length" sheet from Excel
            wire_length_sheet = self.get_sheet_by_name("Wires Lengths")
            if not wire_length_sheet:
                validation_results["wire_length_validation"].append(
                    "'Wires Length' sheet not found in Excel"
                )
            else:
                headers = wire_length_sheet["headers"]
                data_rows = wire_length_sheet["data"]

                # Get column indices
                try:
                    wire_nr_idx = headers.index("Wire Nr") if "Wire Nr" in headers else None
                    length_idx = headers.index("Length") if "Length" in headers else None

                    if None in [wire_nr_idx, length_idx]:
                        missing = [h for h, idx in zip(["Wire Nr", "Length"], 
                                                    [wire_nr_idx, length_idx]) if idx is None]
                        validation_results["wire_length_validation"].append(
                            f"Required columns not found in Excel: {', '.join(missing)}"
                        )
                    else:
                        # Create mapping of wire numbers to lengths from Excel
                        excel_wire_lengths = {}
                        for row in data_rows:
                            wire_nr = row[wire_nr_idx]
                            if wire_nr:
                                try:
                                    excel_wire_lengths[wire_nr.strip()] = float(row[length_idx])
                                except (ValueError, TypeError):
                                    validation_results["wire_length_validation"].append(
                                        f"Invalid length value for wire {wire_nr} in Excel"
                                    )

                        # Get all subassemblies from XML
                        subassemblies = self.get_subassemblies(flatten_attributes=True)
                        
                        # Find wire subassemblies and validate lengths
                        for sub in subassemblies:
                            name = sub.get('name', '')
                            quantity = sub.get('quantity')
                            unit = sub.get('unitofmeasure', '')
                            
                            # Check if this is a wire subassembly (name contains wire pattern like "45(2)")
                            wire_match = re.search(r"(\d+\(\d+\))\s+CUT\b", name, re.IGNORECASE)
                            if wire_match and quantity and unit.lower() in ['per length', 'length']:
                                wire_id = wire_match.group(1)
                                
                                try:
                                    xml_length = float(quantity)
                                except (ValueError, TypeError):
                                    validation_results["wire_length_validation"].append(
                                        f"Invalid quantity value for wire {wire_id} in XML"
                                    )
                                    continue
                                
                                if wire_id in excel_wire_lengths:
                                    excel_length = excel_wire_lengths[wire_id]
                                    if not self._approx_equal(xml_length, excel_length):
                                        validation_results["wire_length_validation"].append(
                                            f"Wire length mismatch for {wire_id}: "
                                            f"XML={xml_length}, Excel={excel_length}"
                                        )
                                else:
                                    validation_results["wire_length_validation"].append(
                                        f"Wire {wire_id} not found in Excel Wires Length sheet"
                                    )

                except Exception as e:
                    validation_results["wire_length_validation"].append(
                        f"Error during wire length validation: {str(e)}"
                    )

        except Exception as e:
            validation_results["wire_length_validation"].append(
                f"Error during wire length validation: {str(e)}"
            )

        # Update overall status if there are any validation issues
        if (validation_results["mismatches"] or 
            validation_results["workcenter_validation"] or 
            validation_results["wire_length_validation"]):
            validation_results["status"] = "fail"
            validation_results["message"] = "Validation completed with mismatches"

        return validation_results

    def generate_report(self, validation_results):
        """
        Generate a formal PDF report with validation results using a professional blue theme.
        Includes wire length validation section.
        """
        # Create PDF object with professional settings
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Set document properties
        pdf.set_title("SBOM Validation Report")
        pdf.set_author("Automated Validation System")
        
        # =============================================
        # Header with Blue Theme
        # =============================================
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 15, "SBOM VALIDATION REPORT", ln=1, align='C', fill=True)
        
        # =============================================
        # Metadata Section
        # =============================================
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Validation Parameters", ln=1, fill=True)
        
        # Create a table for parameters
        pdf.set_font("Arial", '', 10)
        col_widths = [70, 120]
        
        if hasattr(self, 'wcpr'):
            pdf.cell(col_widths[0], 6, "Workcenter Plant Reference:", border=1)
            pdf.cell(col_widths[1], 6, str(self.wcpr), border=1, ln=1)
        
        if hasattr(self, 'wcpar'):
            pdf.cell(col_widths[0], 6, "Workcenter Production Area Reference:", border=1)
            pdf.cell(col_widths[1], 6, str(self.wcpar), border=1, ln=1)
        
        if hasattr(self, 'wcusfa'):
            pdf.cell(col_widths[0], 6, "Workcenter Use Single Final Assembly:", border=1)
            pdf.cell(col_widths[1], 6, str(self.wcusfa), border=1, ln=1)
        
        # Add timestamp
        pdf.ln(5)
        pdf.set_font("Arial", 'I', 8)
        pdf.cell(0, 5, f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1)
        pdf.ln(10)
        
        # =============================================
        # Work Center Validation Results
        # =============================================
        if validation_results.get("workcenter_validation"):
            pdf.set_fill_color(31, 73, 125)  # Dark blue
            pdf.set_text_color(255, 255, 255)  # White
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 8, "Work Center Validation Results", ln=1, fill=True)
            
            # Create comparison table
            pdf.set_fill_color(221, 235, 247)  # Light blue
            pdf.set_text_color(0, 0, 0)  # Black
            pdf.set_font("Arial", 'B', 10)
            
            # Table header
            pdf.cell(60, 7, "Validation Item", border=1, fill=True)
            pdf.cell(40, 7, "Input Value", border=1, fill=True)
            pdf.cell(40, 7, "XML Value", border=1, fill=True)
            pdf.cell(50, 7, "Status", border=1, fill=True, ln=1)
            
            pdf.set_font("Arial", '', 9)
            
            for result in validation_results["workcenter_validation"]:
                if "mismatch" in result.lower():
                    status = "Mismatch"
                    pdf.set_text_color(255, 0, 0)  # Red
                elif "missing" in result.lower():
                    status = "Missing"
                    pdf.set_text_color(255, 165, 0)  # Orange
                else:
                    status = "Info"
                    pdf.set_text_color(0, 0, 0)  # Black
                
                # Extract components from the result string
                if ":" in result:
                    item = result.split(":")[0].strip()
                    values = result.split(":")[1].strip()
                    if "," in values:
                        input_val = values.split(",")[0].split("=")[1].strip()
                        xml_val = values.split(",")[1].split("=")[1].strip()
                    else:
                        input_val = "N/A"
                        xml_val = values.strip()
                else:
                    item = result
                    input_val = "N/A"
                    xml_val = "N/A"
                
                pdf.cell(60, 6, item, border=1)
                pdf.cell(40, 6, input_val, border=1)
                pdf.cell(40, 6, xml_val, border=1)
                pdf.cell(50, 6, status, border=1, ln=1)
                
            pdf.ln(5)
            pdf.set_text_color(0, 0, 0)  # Reset to black
        
        # =============================================
        # Twisted Wires Validation Results
        # =============================================
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Twisted Wires Validation Results", ln=1, fill=True)
        
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        
        if validation_results['status'] == 'success':
            pdf.set_text_color(0, 128, 0)  # Green
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 8, "SUCCESS: All values matched between SBOM and Excel", ln=1)
        else:
            if validation_results['status'] == 'error':
                pdf.set_text_color(255, 0, 0)  # Red
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 8, f"ERROR: {validation_results['message']}", ln=1)
            else:  # 'fail'
                pdf.set_text_color(255, 0, 0)  # Red
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 8, "FAILURE: Mismatches found between SBOM and Excel", ln=1)
                pdf.ln(5)
                
                # Create table for mismatch details
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 6, "Mismatch Details:", ln=1)
                
                pdf.set_font("Arial", '', 8)
                for mismatch in validation_results['mismatches']:
                    # Format mismatch details with bullet points
                    pdf.cell(10, 5, "")
                    pdf.multi_cell(0, 5, f"* {mismatch}")
                    pdf.ln(1)
        
        # =============================================
        # Wire Length Validation Results
        # =============================================
        pdf.ln(10)
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Wire Length Validation Results", ln=1, fill=True)
        
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        
        if not validation_results.get("wire_length_validation"):
            pdf.set_text_color(0, 128, 0)  # Green
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 8, "SUCCESS: All wire lengths matched between SBOM and Excel", ln=1)
        else:
            pdf.set_text_color(255, 0, 0)  # Red
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 8, "FAILURE: Wire length mismatches found", ln=1)
            pdf.ln(5)
            
            # Create table for mismatch details
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 6, "Wire Length Mismatch Details:", ln=1)
            
            pdf.set_font("Arial", '', 8)
            for mismatch in validation_results["wire_length_validation"]:
                # Format mismatch details with bullet points
                pdf.cell(10, 5, "")
                pdf.multi_cell(0, 5, f"* {mismatch}")
                pdf.ln(1)
        
        # =============================================
        # Summary Section
        # =============================================
        pdf.ln(10)
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Validation Summary", ln=1, fill=True)
        
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        
        workcenter_issues = len(validation_results.get("workcenter_validation", []))
        wire_mismatches = len(validation_results.get("mismatches", []))
        wire_length_issues = len(validation_results.get("wire_length_validation", []))
        total_issues = workcenter_issues + wire_mismatches + wire_length_issues
        
        # Create summary table
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(100, 7, "Validation Category", border=1, fill=True)
        pdf.cell(40, 7, "Issues Found", border=1, fill=True)
        pdf.cell(50, 7, "Status", border=1, fill=True, ln=1)
        
        pdf.set_font("Arial", '', 9)
        
        # Work Center row
        status_color = (255, 0, 0) if workcenter_issues > 0 else (0, 128, 0)
        status_text = "FAIL" if workcenter_issues > 0 else "PASS"
        pdf.set_text_color(0, 0, 0)
        pdf.cell(100, 6, "Work Center Attributes", border=1)
        pdf.cell(40, 6, str(workcenter_issues), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Twisted Wires row
        status_color = (255, 0, 0) if wire_mismatches > 0 else (0, 128, 0)
        status_text = "FAIL" if wire_mismatches > 0 else "PASS"
        pdf.set_text_color(0, 0, 0)
        pdf.cell(100, 6, "Twisted Wires Validation", border=1)
        pdf.cell(40, 6, str(wire_mismatches), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Wire Length row
        status_color = (255, 0, 0) if wire_length_issues > 0 else (0, 128, 0)
        status_text = "FAIL" if wire_length_issues > 0 else "PASS"
        pdf.set_text_color(0, 0, 0)
        pdf.cell(100, 6, "Wire Length Validation", border=1)
        pdf.cell(40, 6, str(wire_length_issues), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Overall Status row
        status_color = (255, 0, 0) if total_issues > 0 else (0, 128, 0)
        status_text = "FAIL" if total_issues > 0 else "PASS"
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(100, 6, "OVERALL VALIDATION STATUS", border=1)
        pdf.cell(40, 6, str(total_issues), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Reset text color
        pdf.set_text_color(0, 0, 0)
        
        # =============================================
        # Footer
        # =============================================
        pdf.ln(15)
        pdf.set_font("Arial", 'I', 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, "This is an automatically generated validation report", ln=1, align='C')
        pdf.cell(0, 5, "For any discrepancies, please contact the validation team", ln=1, align='C')
        
        # Save the PDF to a temporary file
        report_dir = "temp_reports"
        os.makedirs(report_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(report_dir, f"sbom_validation_report_{timestamp}.pdf")
        pdf.output(report_path)
        
        return {'report_path': report_path}



    def validate_with_nlp(self):
        if not self.xml_data:
            raise ValueError("No XML data loaded for validation")
        if not self.excel_data:
            raise ValueError("No Excel data loaded for validation")

        validation_results = {
            "status": "success",
            "message": "All values matched.",
            "mismatches": [],
            "workcenter_validation": [],
            "wire_length_validation": [],
            "nlp_processing_notes": []
        }

        # ====================================================================
        # 1. Validate work center attributes from XML against input values
        # ====================================================================
        try:
            # Get the SBOM attributes from the XML data structure
            sbom_attributes = self.get_sbom_attributes()
            if not sbom_attributes:
                raise ValueError("No SBOM attributes found in XML data")
            
            # We'll use the first SBOM's attributes for validation
            attributes = sbom_attributes[0]
            
            # Get the work center attributes
            xml_wcpr = attributes.get('workcenterplantreference')
            xml_wcpar = attributes.get('workcenterproductionareareference')
            xml_wcusfa = attributes.get('workcenter_usesinglefinalassembly')        

            # Check if we have input values to compare against
            if hasattr(self, 'wcpr'):
                if xml_wcpr is None:
                    validation_results["workcenter_validation"].append(
                        "Workcenter Plant Reference missing in XML"
                    )
                elif str(self.wcpr) != str(xml_wcpr):
                    validation_results["workcenter_validation"].append(
                        f"Workcenter Plant Reference mismatch: Input={self.wcpr}, XML={xml_wcpr}"
                    )

            if hasattr(self, 'wcpar'):
                if xml_wcpar is None:
                    validation_results["workcenter_validation"].append(
                        "Workcenter Production Area Reference missing in XML"
                    )
                elif str(self.wcpar) != str(xml_wcpar):
                    validation_results["workcenter_validation"].append(
                        f"Workcenter Production Area Reference mismatch: Input={self.wcpar}, XML={xml_wcpar}"
                    )

            if hasattr(self, 'wcusfa'):
                if xml_wcusfa is None:
                    validation_results["workcenter_validation"].append(
                        "Workcenter Use Single Final Assembly missing in XML"
                    )
                else:
                    # Handle case sensitivity and potential typo in attribute name
                    input_value = str(self.wcusfa).strip().lower()
                    xml_value = str(xml_wcusfa).strip().lower()
                    
                    if input_value != xml_value:
                        validation_results["workcenter_validation"].append(
                            f"Workcenter Use Single Final Assembly mismatch: Input={self.wcusfa}, XML={xml_wcusfa}"
                        )

        except Exception as e:
            validation_results["workcenter_validation"].append(
                f"Error validating work center attributes: {str(e)}"
            )

        # ====================================================================
        # 2. Enhanced NLP-based twisted wires validation
        # ====================================================================
        try:
            cost_results = self.get_cost_results()
            if not cost_results:
                validation_results["status"] = "error"
                validation_results["message"] = "No cost results found in SBOM XML."
                return validation_results

            # Validate against Excel sheet "Twisted Wires"
            twisted_sheet = self.get_sheet_by_name("Twisted Wires")
            if not twisted_sheet:
                validation_results["status"] = "error"
                validation_results["message"] = "'Twisted Wires' sheet not found in Excel."
                return validation_results

            headers = twisted_sheet["headers"]
            data_rows = twisted_sheet["data"]

            # Get column indices for all relevant fields
            try:
                wire_nr_idx = headers.index("Wires Nr") if "Wires Nr" in headers else None
                pitch_idx = headers.index("Pitch") if "Pitch" in headers else None
                open_end1_idx = headers.index("Open end Length 1") if "Open end Length 1" in headers else None
                open_end2_idx = headers.index("Open end Length 2") if "Open end Length 2" in headers else None
                twist_len_idx = headers.index("Length of twist") if "Length of twist" in headers else None

                if None in [wire_nr_idx, pitch_idx, open_end1_idx, open_end2_idx, twist_len_idx]:
                    missing = [h for h, idx in zip(["Wires Nr", "Pitch", "Open end Length 1", 
                                                "Open end Length 2", "Length of twist"], 
                                                [wire_nr_idx, pitch_idx, open_end1_idx, 
                                                open_end2_idx, twist_len_idx]) if idx is None]
                    validation_results["status"] = "error"
                    validation_results["message"] = f"Required columns not found in Excel: {', '.join(missing)}"
                    return validation_results

                # Create mapping of wire numbers to all relevant values from Excel
                excel_wire_data = {}
                for row in data_rows:
                    wire_nr = row[wire_nr_idx]
                    if wire_nr:
                        clean_wire_nr = re.sub(r'[^0-9()]', '', str(wire_nr))  # Clean special chars
                        excel_wire_data[clean_wire_nr] = {
                            "pitch": row[pitch_idx],
                            "open_end1": row[open_end1_idx],
                            "open_end2": row[open_end2_idx],
                            "twist_len": row[twist_len_idx]
                        }

                # Process each cost result with NLP
                for result in cost_results:
                    description = result.get("description", "")
                    if not description:
                        continue
                        
                    wire_info = self._extract_wire_info_nlp(description)
                    validation_results["nlp_processing_notes"].append(
                        f"Processed description: {description[:50]}... "
                        f"(Confidence: {wire_info['confidence']:.1f})"
                    )

                    # Validate each wire in this cost result against Excel
                    for wire in wire_info['wires']:
                        clean_wire = re.sub(r'[^0-9()]', '', wire)  # Clean special chars
                        
                        if clean_wire not in excel_wire_data:
                            validation_results["mismatches"].append(f"{clean_wire} missing in Excel")
                            continue
                            
                        excel_data = excel_wire_data[clean_wire]
                        
                        # Validate each field with tolerance
                        self._validate_nlp_field(
                            validation_results, 
                            "pitch", 
                            wire_info.get('pitch'), 
                            excel_data.get("pitch"), 
                            clean_wire
                        )
                        
                        self._validate_nlp_field(
                            validation_results,
                            "open end length 1 (untwist A)",
                            wire_info.get('untwist_a'),
                            excel_data.get("open_end1"),
                            clean_wire
                        )
                        
                        self._validate_nlp_field(
                            validation_results,
                            "open end length 2 (untwist B)",
                            wire_info.get('untwist_b'),
                            excel_data.get("open_end2"),
                            clean_wire
                        )
                        
                        self._validate_nlp_field(
                            validation_results,
                            "twist length",
                            wire_info.get('twist_length'),
                            excel_data.get("twist_len"),
                            clean_wire
                        )

            except Exception as e:
                validation_results["status"] = "error"
                validation_results["message"] = f"Error during twisted wires validation: {str(e)}"
                return validation_results

        except Exception as e:
            validation_results["status"] = "error"
            validation_results["message"] = f"Error during validation: {str(e)}"
            return validation_results

        # ====================================================================
        # 3. Validate wire lengths from XML subassemblies against Excel
        # (Same implementation as standard validate() method)
        # ====================================================================
        try:
            # Get the "Wires Length" sheet from Excel
            wire_length_sheet = self.get_sheet_by_name("Wires Lengths")
            if not wire_length_sheet:
                validation_results["wire_length_validation"].append(
                    "'Wires Lengths' sheet not found in Excel"
                )
            else:
                headers = wire_length_sheet["headers"]
                data_rows = wire_length_sheet["data"]

                # Get column indices
                try:
                    wire_nr_idx = headers.index("Wire Nr") if "Wire Nr" in headers else None
                    length_idx = headers.index("Length") if "Length" in headers else None

                    if None in [wire_nr_idx, length_idx]:
                        missing = [h for h, idx in zip(["Wire Nr", "Length"], 
                                                    [wire_nr_idx, length_idx]) if idx is None]
                        validation_results["wire_length_validation"].append(
                            f"Required columns not found in Excel: {', '.join(missing)}"
                        )
                    else:
                        # Create mapping of wire numbers to lengths from Excel
                        excel_wire_lengths = {}
                        for row in data_rows:
                            wire_nr = row[wire_nr_idx]
                            if wire_nr:
                                try:
                                    excel_wire_lengths[wire_nr.strip()] = float(row[length_idx])
                                except (ValueError, TypeError):
                                    validation_results["wire_length_validation"].append(
                                        f"Invalid length value for wire {wire_nr} in Excel"
                                    )

                        # Get all subassemblies from XML
                        subassemblies = self.get_subassemblies(flatten_attributes=True)
                        
                        # Find wire subassemblies and validate lengths
                        for sub in subassemblies:
                            name = sub.get('name', '')
                            quantity = sub.get('quantity')
                            unit = sub.get('unitofmeasure', '')
                            
                            # Check if this is a wire subassembly (name contains wire pattern like "45(2)")
                            wire_match = re.search(r"(\d+\(\d+\))\s+CUT\b", name, re.IGNORECASE)
                            if wire_match and quantity and unit.lower() in ['per length', 'length']:
                                wire_id = wire_match.group(1)
                                
                                try:
                                    xml_length = float(quantity)
                                except (ValueError, TypeError):
                                    validation_results["wire_length_validation"].append(
                                        f"Invalid quantity value for wire {wire_id} in XML"
                                    )
                                    continue
                                
                                if wire_id in excel_wire_lengths:
                                    excel_length = excel_wire_lengths[wire_id]
                                    if not self._approx_equal(xml_length, excel_length):
                                        validation_results["wire_length_validation"].append(
                                            f"Wire length mismatch for {wire_id}: "
                                            f"XML={xml_length}, Excel={excel_length}"
                                        )
                                else:
                                    validation_results["wire_length_validation"].append(
                                        f"Wire {wire_id} not found in Excel Wires Length sheet"
                                    )

                except Exception as e:
                    validation_results["wire_length_validation"].append(
                        f"Error during wire length validation: {str(e)}"
                    )

        except Exception as e:
            validation_results["wire_length_validation"].append(
                f"Error during wire length validation: {str(e)}"
            )

        # Update overall status if there are any validation issues
        if (validation_results["mismatches"] or 
            validation_results["workcenter_validation"] or 
            validation_results["wire_length_validation"]):
            validation_results["status"] = "fail"
            validation_results["message"] = "Validation completed with mismatches"

        return validation_results

    def _extract_wire_info_nlp(self, description):
        """Enhanced NLP method to extract wire information with better accuracy."""
        doc = self.nlp(description.lower())  # Process in lowercase for consistency
        
        # Initialize result dict with all possible fields
        result = {
            'wires': [],
            'pitch': None,
            'untwist_a': None,
            'untwist_b': None,
            'twist_length': None,
            'direction': None,
            'colors': [],
            'confidence': 1.0  # Track confidence in extraction
        }
        
        # First pass: Identify wire numbers and potential colors
        wire_pattern = re.compile(r'(\d+\(\d+\))')
        wires_found = wire_pattern.findall(description)
        
        if not wires_found:
            # Try more flexible pattern if strict pattern fails
            wires_found = re.findall(r'(\d+\s*\(\s*\d+\s*\))', description)
        
        # Clean up wire numbers
        result['wires'] = [w.strip() for w in wires_found]
        
        # Extract colors if present (common wire colors)
        color_keywords = ['wh', 'bu', 'gy', 'bk', 'rd', 'ye', 'gn', 'bn']
        color_matches = []
        for token in doc:
            if token.text.lower() in color_keywords:
                color_matches.append(token.text.upper())
        result['colors'] = color_matches
        
        # Extract numerical values with context awareness
        current_measurement = None
        
        for token in doc:
            # Look for measurement types
            if token.text in ['pitch', 'untwist', 'length', 'twist']:
                current_measurement = token.text
                continue
                
            # Look for direction indicators
            if token.text in ['s', 'z']:
                result['direction'] = token.text.upper()
                continue
                
            # Process numbers with measurement context
            if token.like_num and current_measurement:
                try:
                    value = float(token.text)
                    
                    if current_measurement == 'pitch':
                        result['pitch'] = value
                    elif current_measurement == 'untwist':
                        # Look ahead for a/b indicators
                        next_tokens = [t.text for t in doc[token.i:token.i+3]]
                        if 'a' in next_tokens:
                            result['untwist_a'] = value
                        elif 'b' in next_tokens:
                            result['untwist_b'] = value
                    elif current_measurement in ['length', 'twist']:
                        # Check if this is twist length
                        if 'twist' in [t.text for t in token.head.children]:
                            result['twist_length'] = value
                            
                    current_measurement = None  # Reset after processing
                except ValueError:
                    continue
        
        # Fallback: If no structured data found, try pattern matching
        if not any([result['pitch'], result['untwist_a'], result['untwist_b'], result['twist_length']]):
            # Try to extract from common patterns
            pitch_match = re.search(r'pitch:\s*([\d.]+)', description.lower())
            if pitch_match:
                result['pitch'] = float(pitch_match.group(1))
                
            untwist_a_match = re.search(r'untwist a:\s*([\d.]+)', description.lower())
            if untwist_a_match:
                result['untwist_a'] = float(untwist_a_match.group(1))
                
            untwist_b_match = re.search(r'untwist b:\s*([\d.]+)', description.lower())
            if untwist_b_match:
                result['untwist_b'] = float(untwist_b_match.group(1))
                
            twist_match = re.search(r'twist length:\s*([\d.]+)', description.lower())
            if twist_match:
                result['twist_length'] = float(twist_match.group(1))
            
            # Lower confidence if we had to use fallback
            result['confidence'] = 0.7
        
        return result

    def _validate_nlp_field(self, validation_results, field_name, xml_value, excel_value, wire_id):
        """Helper method to validate a single field with proper error messaging."""
        if xml_value is None:
            validation_results["nlp_processing_notes"].append(
                f"Could not extract {field_name} for wire {wire_id} from XML description"
            )
            return
            
        if excel_value is None:
            validation_results["mismatches"].append(
                f"{field_name.capitalize()} missing in Excel for wire {wire_id}"
            )
            return
            
        try:
            excel_num = float(excel_value)
            if not self._approx_equal(xml_value, excel_num):
                validation_results["mismatches"].append(
                    f"{field_name.capitalize()} mismatch for {wire_id}: "
                    f"XML={xml_value}, Excel={excel_num}"
                )
        except (ValueError, TypeError):
            validation_results["mismatches"].append(
                f"Invalid {field_name} value in Excel for wire {wire_id}: {excel_value}"
            )

    def _approx_equal(self, a, b, tolerance=0.01):
        """Helper method to compare floating point numbers with tolerance."""
        return abs(a - b) <= tolerance

    def generate_nlp_report(self, validation_results):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        pdf.set_title("SBOM NLP Validation Report")
        pdf.set_author("Automated NLP Validation System")
        pdf.set_fill_color(31, 73, 125)
        pdf.set_text_color(255, 255, 255)  
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 15, "SBOM NLP VALIDATION REPORT", ln=1, align='C', fill=True)
        
        pdf.set_fill_color(221, 235, 247)  
        pdf.set_text_color(0, 0, 0) 
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "Validation Parameters", ln=1, fill=True)
        
        # Create a table for parameters
        pdf.set_font("Arial", '', 10)
        col_widths = [70, 120]
        
        if hasattr(self, 'wcpr'):
            pdf.cell(col_widths[0], 6, "Workcenter Plant Reference:", border=1)
            pdf.cell(col_widths[1], 6, str(self.wcpr), border=1, ln=1)
        
        if hasattr(self, 'wcpar'):
            pdf.cell(col_widths[0], 6, "Workcenter Production Area Reference:", border=1)
            pdf.cell(col_widths[1], 6, str(self.wcpar), border=1, ln=1)
        
        if hasattr(self, 'wcusfa'):
            pdf.cell(col_widths[0], 6, "Workcenter Use Single Final Assembly:", border=1)
            pdf.cell(col_widths[1], 6, str(self.wcusfa), border=1, ln=1)
        
        # Add NLP information
        pdf.cell(col_widths[0], 6, "NLP Model Used:", border=1)
        pdf.cell(col_widths[1], 6, "en_core_web_sm", border=1, ln=1)
        
        # Add timestamp
        pdf.ln(5)
        pdf.set_font("Arial", 'I', 8)
        pdf.cell(0, 5, f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1)
        pdf.ln(10)
        
        # =============================================
        # Work Center Validation Results
        # =============================================
        if validation_results.get("workcenter_validation"):
            pdf.set_fill_color(31, 73, 125)  # Dark blue
            pdf.set_text_color(255, 255, 255)  # White
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 8, "Work Center Validation Results", ln=1, fill=True)
            
            # Create comparison table
            pdf.set_fill_color(221, 235, 247)  # Light blue
            pdf.set_text_color(0, 0, 0)  # Black
            pdf.set_font("Arial", 'B', 10)
            
            # Table header
            pdf.cell(60, 7, "Validation Item", border=1, fill=True)
            pdf.cell(40, 7, "Input Value", border=1, fill=True)
            pdf.cell(40, 7, "XML Value", border=1, fill=True)
            pdf.cell(50, 7, "Status", border=1, fill=True, ln=1)
            
            pdf.set_font("Arial", '', 9)
            
            for result in validation_results["workcenter_validation"]:
                if "mismatch" in result.lower():
                    status = "Mismatch"
                    pdf.set_text_color(255, 0, 0)  # Red
                elif "missing" in result.lower():
                    status = "Missing"
                    pdf.set_text_color(255, 165, 0)  # Orange
                else:
                    status = "Info"
                    pdf.set_text_color(0, 0, 0)  # Black
                
                # Extract components from the result string
                if ":" in result:
                    item = result.split(":")[0].strip()
                    values = result.split(":")[1].strip()
                    if "," in values:
                        input_val = values.split(",")[0].split("=")[1].strip()
                        xml_val = values.split(",")[1].split("=")[1].strip()
                    else:
                        input_val = "N/A"
                        xml_val = values.strip()
                else:
                    item = result
                    input_val = "N/A"
                    xml_val = "N/A"
                
                pdf.cell(60, 6, item, border=1)
                pdf.cell(40, 6, input_val, border=1)
                pdf.cell(40, 6, xml_val, border=1)
                pdf.cell(50, 6, status, border=1, ln=1)
                
            pdf.ln(5)
            pdf.set_text_color(0, 0, 0)  # Reset to black
        
        # =============================================
        # NLP-based Twisted Wires Validation Results
        # =============================================
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "NLP-based Twisted Wires Validation Results", ln=1, fill=True)
        
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        
        if validation_results['status'] == 'success':
            pdf.set_text_color(0, 128, 0)  # Green
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 8, "SUCCESS: All values matched between SBOM and Excel", ln=1)
        else:
            if validation_results['status'] == 'error':
                pdf.set_text_color(255, 0, 0)  # Red
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 8, f"ERROR: {validation_results['message']}", ln=1)
            else:  # 'fail'
                pdf.set_text_color(255, 0, 0)  # Red
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 8, "FAILURE: Mismatches found between SBOM and Excel", ln=1)
                pdf.ln(5)
                
                # Create table for mismatch details
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(0, 6, "NLP Processing Details:", ln=1)
                
                pdf.set_font("Arial", '', 8)
                for mismatch in validation_results['mismatches']:
                    # Format mismatch details with bullet points
                    pdf.cell(10, 5, "")
                    pdf.multi_cell(0, 5, f"* {mismatch}")
                    pdf.ln(1)
        
        # =============================================
        # Summary Section
        # =============================================
        pdf.ln(10)
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "NLP Validation Summary", ln=1, fill=True)
        
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        
        workcenter_issues = len(validation_results.get("workcenter_validation", []))
        wire_mismatches = len(validation_results.get("mismatches", []))
        total_issues = workcenter_issues + wire_mismatches
        
        # Create summary table
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(100, 7, "Validation Category", border=1, fill=True)
        pdf.cell(40, 7, "Issues Found", border=1, fill=True)
        pdf.cell(50, 7, "Status", border=1, fill=True, ln=1)
        
        pdf.set_font("Arial", '', 9)
        
        # Work Center row
        status_color = (255, 0, 0) if workcenter_issues > 0 else (0, 128, 0)
        status_text = "FAIL" if workcenter_issues > 0 else "PASS"
        pdf.set_text_color(0, 0, 0)
        pdf.cell(100, 6, "Work Center Attributes", border=1)
        pdf.cell(40, 6, str(workcenter_issues), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Twisted Wires (NLP) row
        status_color = (255, 0, 0) if wire_mismatches > 0 else (0, 128, 0)
        status_text = "FAIL" if wire_mismatches > 0 else "PASS"
        pdf.set_text_color(0, 0, 0)
        pdf.cell(100, 6, "Twisted Wires (NLP Validation)", border=1)
        pdf.cell(40, 6, str(wire_mismatches), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Overall Status row
        status_color = (255, 0, 0) if total_issues > 0 else (0, 128, 0)
        status_text = "FAIL" if total_issues > 0 else "PASS"
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(100, 6, "OVERALL NLP VALIDATION STATUS", border=1)
        pdf.cell(40, 6, str(total_issues), border=1)
        pdf.set_text_color(*status_color)
        pdf.cell(50, 6, status_text, border=1, ln=1)
        
        # Reset text color
        pdf.set_text_color(0, 0, 0)
        
        # =============================================
        # NLP Processing Notes
        # =============================================
        pdf.ln(10)
        pdf.set_fill_color(31, 73, 125)  # Dark blue
        pdf.set_text_color(255, 255, 255)  # White
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(0, 8, "NLP Processing Notes", ln=1, fill=True)
        
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_text_color(0, 0, 0)  # Black
        pdf.set_font("Arial", '', 10)
        
        notes = [
            "1. This validation used spaCy's NLP model (en_core_web_sm) to parse wire descriptions.",
            "2. The system extracts wire numbers, colors, and measurements from natural language descriptions.",
            "3. Comparisons are made with a tolerance of ±0.01 for floating point values.",
            "4. Some descriptions may be parsed differently than regex-based approaches."
        ]
        
        for note in notes:
            pdf.cell(10, 6, "")
            pdf.multi_cell(0, 6, note)
            pdf.ln(2)
        
        # =============================================
        # Footer
        # =============================================
        pdf.ln(15)
        pdf.set_font("Arial", 'I', 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, "This is an automatically generated NLP validation report", ln=1, align='C')
        pdf.cell(0, 5, "For any discrepancies, please contact the validation team", ln=1, align='C')
        
        # Save the PDF to a temporary file
        report_dir = "temp_reports"
        os.makedirs(report_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(report_dir, f"sbom_nlp_validation_report_{timestamp}.pdf")
        pdf.output(report_path)
        
        return {'report_path': report_path}

    def _approx_equal(self, a, b, tolerance=0.01):
        """Helper method to compare floating point numbers with tolerance."""
        return abs(a - b) <= tolerance